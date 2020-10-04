import openpyxl

wb = openpyxl.load_workbook("C://Users//hannez//Desktop//pyxl.xlsx", data_only=True) #parses the actual values and not the formulas

ws = wb["pyxl"] #set the current worksheet to work with
# first column = participantcodes
# second column = chosen lottery

lottery_choices_dict = {}

for col in ws.iter_cols(min_row=2, min_col=1, max_col=1, max_row=ws.max_row): #iterate through isin column ws.max_row
    for cell in col:
        currentrow = cell.row
        # print(cell.value)
        # print(ws.cell(row = currentrow, column = 2).value)
        if cell.value in lottery_choices_dict:
            print(f"{cell.value} already in dict")
        else:
            lottery_choices_dict[cell.value] = ws.cell(row = currentrow, column = 2).value

print(lottery_choices_dict)
print(len(lottery_choices_dict))

# wb.save("C://Users//hannez//Desktop//pyxl.xlsx")